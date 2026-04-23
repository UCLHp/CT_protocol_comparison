import pandas as pd
from bs4 import BeautifulSoup
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

unique_id_cols = ["Protocol", "Acquisition Number", "Label", "Type", "Result Label"]

gray_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")      # removed
green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")    # added
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")   # changed cell
row_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")      # changed row


def _center_window(win, width, height):
    win.update_idletasks()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    x = (sw - width) // 2
    y = (sh - height) // 2
    win.geometry(f"{width}x{height}+{x}+{y}")


def choose_mode_and_files(root):
    choice = {"mode": None}

    dialog = tk.Toplevel(root)
    dialog.title("CT Protocol Tool")
    dialog.resizable(False, False)

    dialog.attributes("-topmost", True)
    _center_window(dialog, 520, 170)

    def set_mode(mode):
        choice["mode"] = mode
        dialog.destroy()

    def cancel():
        choice["mode"] = None
        dialog.destroy()

    dialog.protocol("WM_DELETE_WINDOW", cancel)
    dialog.bind("<Escape>", lambda _e: cancel())

    tk.Label(dialog, text="Choose mode:").pack(padx=16, pady=(16, 10))

    frame = tk.Frame(dialog)
    frame.pack(padx=16, pady=(0, 10))

    tk.Button(
        frame,
        text="1 file: Extract",
        width=22,
        command=lambda: set_mode("extract")
    ).grid(row=0, column=0, padx=6, pady=4)

    tk.Button(
        frame,
        text="2 files: Compare",
        width=22,
        command=lambda: set_mode("compare")
    ).grid(row=0, column=1, padx=6, pady=4)

    tk.Button(dialog, text="Cancel", width=12, command=cancel).pack(pady=(0, 14))

    dialog.lift()
    dialog.focus_force()
    dialog.after(300, lambda: dialog.attributes("-topmost", False))

    root.wait_window(dialog)

    mode = choice["mode"]

    if mode is None:
        return []

    if mode == "extract":
        path = filedialog.askopenfilename(
            parent=root,
            title="Select ONE HTML file to extract",
            filetypes=[("HTML files", "*.html")]
        )
        return [path] if path else []

    before_path = filedialog.askopenfilename(
        parent=root,
        title="Select BEFORE (older) HTML file",
        filetypes=[("HTML files", "*.html")]
    )

    if not before_path:
        return []

    after_path = filedialog.askopenfilename(
        parent=root,
        title="Select AFTER (newer) HTML file",
        filetypes=[("HTML files", "*.html")]
    )

    if not after_path:
        return []

    return [before_path, after_path]


def select_output_folder(root):
    return filedialog.askdirectory(parent=root, title="Select Folder to Save Outputs")


def parse_html_file(html_path):
    with open(html_path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    combined_data = []
    current_protocol = None
    acquisition_number = None
    current_acquisition = None
    current_result = None
    in_result_section = False

    for element in soup.find_all(["p", "table"]):

        if element.name == "p" and "Acquisition label" in element.get_text():
            prev_exam = element.find_previous("p", class_="exam")
            current_protocol = prev_exam.get_text(strip=True) if prev_exam else ""

            acq_text = element.get_text(strip=True)
            acquisition_number = acq_text.split(",")[0].replace("Acquisition label :", "").strip()
            label = acq_text.split(",")[1].strip() if "," in acq_text else ""

            current_acquisition = {
                "Protocol": current_protocol,
                "Acquisition Number": acquisition_number,
                "Label": label,
                "Type": "Acquisition",
                "Result Label": ""
            }

            combined_data.append(current_acquisition)
            in_result_section = False

        elif element.name == "table" and current_acquisition and not in_result_section:
            for row in element.find_all("tr"):
                cells = row.find_all("td")

                if len(cells) == 2:
                    parameter = cells[0].get_text(strip=True)
                    value = cells[1].get_text(strip=True)
                    current_acquisition[parameter] = value

        elif element.name == "p" and "Result Label" in element.get_text():
            result_label = element.get_text(strip=True).replace("Result Label :", "").strip()

            current_result = {
                "Protocol": current_protocol,
                "Acquisition Number": acquisition_number,
                "Label": "",
                "Type": "Result",
                "Result Label": result_label
            }

            combined_data.append(current_result)
            in_result_section = True

        elif element.name == "table" and in_result_section:
            for row in element.find_all("tr"):
                cells = row.find_all("td")

                if len(cells) == 2:
                    parameter = cells[0].get_text(strip=True)
                    value = cells[1].get_text(strip=True)
                    current_result[parameter] = value

    df = pd.DataFrame(combined_data).fillna("")
    df.columns = df.columns.str.strip()

    df["Acquisition Number"] = pd.to_numeric(df["Acquisition Number"], errors="coerce")

    df.sort_values(
        by=["Protocol", "Acquisition Number", "Type", "Result Label"],
        inplace=True
    )

    return df


def df_to_dict_list(df):
    result = {}

    for _, row in df.iterrows():
        key = tuple(str(row.get(col, "")).strip() for col in unique_id_cols)
        result.setdefault(key, []).append(row.to_dict())

    return result


def compare_files(df_before, df_after):
    dict_before = df_to_dict_list(df_before)
    dict_after = df_to_dict_list(df_after)

    keys_before = set(dict_before.keys())
    keys_after = set(dict_after.keys())

    removed_keys = keys_before - keys_after
    added_keys = keys_after - keys_before

    removed_rows = []
    added_rows = []
    changed_list = []

    for key in removed_keys:
        removed_rows.extend(dict_before[key])

    for key in added_keys:
        added_rows.extend(dict_after[key])

    # This is the important update:
    # Compare all parameter columns from BOTH before and after files.
    all_param_cols = sorted(
        (set(df_before.columns) | set(df_after.columns)) - set(unique_id_cols)
    )

    for key in keys_before & keys_after:
        rows_before = dict_before[key]
        rows_after = dict_after[key]

        for rb, ra in zip(rows_before, rows_after):
            for col in all_param_cols:
                val_b = str(rb.get(col, "")).strip()
                val_a = str(ra.get(col, "")).strip()

                if val_b != val_a:
                    changed_list.append({
                        **{c: rb.get(c, "") for c in unique_id_cols},
                        "Parameter": col,
                        "Before": val_b,
                        "After": val_a
                    })

    return (
        pd.DataFrame(removed_rows),
        pd.DataFrame(added_rows),
        pd.DataFrame(changed_list)
    )


def align_columns_for_output(df_before, df_after):
    """
    Makes BEFORE and AFTER Excel outputs have the same columns.

    This helps when a parameter/header exists only in the AFTER file
    or only in the BEFORE file.
    """

    all_cols = list(unique_id_cols)

    param_cols = sorted(
        (set(df_before.columns) | set(df_after.columns)) - set(unique_id_cols)
    )

    all_cols.extend(param_cols)

    df_before_aligned = df_before.reindex(columns=all_cols, fill_value="")
    df_after_aligned = df_after.reindex(columns=all_cols, fill_value="")

    return df_before_aligned, df_after_aligned


def highlight_rows(filepath, output_path, highlight_df, fill_color):
    if highlight_df.empty:
        return

    wb = load_workbook(filepath)
    ws = wb.active

    df_excel = pd.read_excel(filepath).fillna("")
    df_excel.columns = df_excel.columns.str.strip()

    highlight_df = highlight_df.copy()
    highlight_df.columns = highlight_df.columns.str.strip()
    highlight_df = highlight_df.drop_duplicates(subset=unique_id_cols)

    merged = df_excel.reset_index().merge(
        highlight_df[unique_id_cols],
        on=unique_id_cols,
        how="inner"
    )

    for row_idx in merged["index"].tolist():
        for cell in ws[row_idx + 2]:
            cell.fill = fill_color

    wb.save(output_path)


def highlight_changes(filepath, output_path, changed_df):
    if changed_df.empty:
        return

    wb = load_workbook(output_path)
    ws = wb.active

    df_excel = pd.read_excel(filepath).fillna("")
    df_excel.columns = df_excel.columns.str.strip()

    changed_df = changed_df.copy()
    changed_df = changed_df.drop_duplicates(subset=unique_id_cols + ["Parameter"])

    merged = df_excel.reset_index().merge(
        changed_df[unique_id_cols + ["Parameter"]],
        on=unique_id_cols,
        how="inner"
    )

    highlighted_rows = set()

    for _, row in merged.iterrows():
        row_idx = int(row["index"]) + 2
        col_name = row["Parameter"]

        if col_name in df_excel.columns:
            col_idx = df_excel.columns.get_loc(col_name) + 1

            # Highlight changed cell orange
            ws.cell(row=row_idx, column=col_idx).fill = orange_fill

            # Highlight rest of changed row pale yellow
            if row_idx not in highlighted_rows:
                for c in ws[row_idx]:
                    if c.fill != orange_fill:
                        c.fill = row_fill

                highlighted_rows.add(row_idx)

    wb.save(output_path)


def autosize_excel_columns(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    wb.save(filepath)


def main():
    root = tk.Tk()
    root.withdraw()

    html_files = choose_mode_and_files(root)

    if not html_files:
        print("Cancelled.")
        raise SystemExit

    save_folder = select_output_folder(root)

    if not save_folder:
        print("Cancelled.")
        raise SystemExit

    if len(html_files) == 1:
        df = parse_html_file(html_files[0])

        out_path = os.path.join(save_folder, "protocol_summary.xlsx")
        df.to_excel(out_path, index=False)
        autosize_excel_columns(out_path)

        print(f"Extracted single file to: {out_path}")
        raise SystemExit

    before_path, after_path = html_files[0], html_files[1]

    df_before = parse_html_file(before_path)
    df_after = parse_html_file(after_path)

    removed_df, added_df, changed_df = compare_files(df_before, df_after)

    report_path = os.path.join(save_folder, "comparison_report.xlsx")

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        removed_df.to_excel(writer, sheet_name="Removed", index=False)
        added_df.to_excel(writer, sheet_name="Added", index=False)
        changed_df.to_excel(writer, sheet_name="Changed", index=False)

        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]

            for column_cells in worksheet.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    before_base = os.path.splitext(os.path.basename(before_path))[0]
    after_base = os.path.splitext(os.path.basename(after_path))[0]

    before_out = os.path.join(save_folder, f"BEFORE_{before_base}_highlighted.xlsx")
    after_out = os.path.join(save_folder, f"AFTER_{after_base}_highlighted.xlsx")

    # Align columns so new/removed parameter headers are visible in both outputs
    df_before_out, df_after_out = align_columns_for_output(df_before, df_after)

    df_before_out.to_excel(before_out, index=False)
    df_after_out.to_excel(after_out, index=False)

    highlight_rows(before_out, before_out, removed_df, gray_fill)
    highlight_rows(after_out, after_out, added_df, green_fill)

    highlight_changes(before_out, before_out, changed_df)
    highlight_changes(after_out, after_out, changed_df)

    autosize_excel_columns(before_out)
    autosize_excel_columns(after_out)

    print(f"Comparison report saved to: {report_path}")
    print(f"Highlighted BEFORE saved to: {before_out}")
    print(f"Highlighted AFTER  saved to: {after_out}")


if __name__ == "__main__":
    main()